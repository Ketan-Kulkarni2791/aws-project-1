import json, csv
import mysql.connector
from openpyxl import Workbook
import io
from openpyxl.styles import Font
import boto3


queue=boto3.client('sqs')
queue_url=queue.get_queue_url(QueueName='final-task-queue-2')['QueueUrl']

def lambda_handler(event, context):
    # TODO implement
    
    repFlg1 = repFlg2 = repFlg3 = emailFlg = "False"
    
    if event['Records'][0]['body'] == "True":
        try:
            
            print("************************** In here")
            
            wb = Workbook()
            
            # Connection with RDS database
            
            my_connection = mysql.connector.connect(
            host="project.c0vrpj3mmoqn.ap-south-1.rds.amazonaws.com",
            user="admin",
            passwd="Pass_123",
            database="sales"
            )
            cursor = my_connection.cursor()
        
            # Generation of report 1
        
            cursor.execute("""Select Region, SUM(Total_Revenue), SUM(Total_Cost), SUM(Total_Profit), Order_Date from sales_copy_1 where 
                              Order_Date LIKE "%-04-2015%" GROUP BY Region;""")
            records_1 = cursor.fetchall()
            records_1_Col = ['Region', 'Total_Revenue', 'Total_Cost', 'Total_Profit', 'Order_Date']
            try:
                ws = wb.create_sheet(0)
                ws.title = "report_1"
                ws.append(records_1_Col)
                for row in records_1:
                    ws.append(list(row))
            
                filepath = "/tmp/"
                wb.save(filepath + "report_1.csv")
                print("------------------- Report 1 saved successfully")
                repFlg1 = "True"
            except:
                print("CSV Report 1 Generation failed")
            
            # Generation of report 2
            
            cursor.execute("""Select Region, SUM(Total_Revenue), SUM(Total_Cost), SUM(Total_Profit), Order_Date from sales_copy_1 where 
                              Order_Date LIKE "%-2015%" GROUP BY Region;""")
            records_2 = cursor.fetchall()
            try:
                ws = wb.create_sheet(0)
                ws.title = "report_2"
                ws.append(records_1_Col)
                for row in records_2:
                    ws.append(list(row))
            
                wb.save(filepath + "report_2.csv")
                print("------------------- Report 2 saved successfully")
                repFlg2 = "True"
            except:
                print("CSV Report 2 Generation failed")
            
            # Generation of report 3
            
            cursor.execute("""SELECT Region, Order_Date, SUM(Units_Sold) from sales_copy_1 WHERE Order_Date LIKE "%-2015%" 
                              and Sales_Channel IN ("Online", "Offline") GROUP BY Region;""")
            records_3 = cursor.fetchall()
            records_3_Col = ['Region', 'Units_Sold', 'Order_Date']
            try:
                ws = wb.create_sheet(0)
                ws.title = "report_3"
                ws.append(records_3_Col)
                for row in records_3:
                    ws.append(list(row))
            
                wb.save(filepath + "report_3.csv")
                print("------------------- Report 3 saved successfully")
                repFlg3 = "True"
            except:
                print("CSV Report 1 Generation failed")
            
            if repFlg1 == "True" and repFlg2 == "True" and repFlg3 == "True":
                emailFlg = "True"
            write_resp=queue.send_message(QueueUrl=queue_url,MessageBody=emailFlg)
            print("written to quere response",write_resp)
            
            
        except mysql.connector.Error as e:  
            print("Here")
            print("Error reading data from MySQL table", e)
        finally:
            if my_connection.is_connected():
                my_connection.close()
                cursor.close()
                print("MySQL connection is closed")
    else:
        return {
            'statusCode': 404,
            'body': json.dumps('Nooo Hello from Lambda!')
        }
